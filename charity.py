import tkinter as tk from tkinter import *

#from openpyxl import load_workbook

path = "D:\data.xlsx" #existing file to work(Not a new file) #wb = load_workbook(filename=path,read_only=False)

from openpyxl import Workbook

wb=Workbook() #create workbook

#ws = wb.active #work on default worksheet ws = wb.create_sheet("register") #create a worksheet

my_w = tk.TK() my_w.geometry("400*250") my_w.title("Charity")

#add one label l10=tk.Label(my_w,text=" Donor Name",font=16,width=30,anchor="c") l10.grid(row=1,column=1,columnspan=4)

l11=tk.Label(my_w,text="Name:",width=10,anchor="c") l11.grid(row=3,column=1)

#add one text box t1=tk.Text(my_w,height=1,width=10,bg="white") t1.grid(row=3,column=2)

l12=tk.Label(my_w,text="occupation:",width=10) l12.grid(row=4,column=1)

#add second text box t2=tk.Label(my_w,height=1,width=4,bg="white") t2.grid(row=4,column=2)

l13=tk.Label(my_w,text="phonenumber",width=10) l13.grid(row=5,column=1)

#add third text box t3=tk.Label(my_w,height=1,width=6,bg="white") t3.grid(row=5,column=2)

radio_v=tk.stringvar() r1=tk.Radiobutton(my_w,text="cash",variable=radio_v,value="cash") r1.grid(row=6,column=1,pady=10)

r2=tk.Radiobutton(my_w,text="cheque",variable=radio_v,value="cheque") r2.grid(row=6,column=2)

r3=tk.Radiobutton(my_w,text="UPI",variable=radio_v,value="UPI") r3.grid(row=6,column=3)

l15=tk.Label(my_w,text="Amount:",width=10,) l15.grid(row=7,column=1)

#add fourth text box t4=tk.Text(my_w,height=1,width=4,bg="white") t4.grid(row=7,column=2)

b1=tk.Button(my_w,text="Add record",width=10,command=lambda : add_data()) b1.grid(row=8,coulmn=2) my_str = tk.stringvar()

l16 = tk.Label(my_w,textvariable=my_str,width=10) l16.grid(row=3,column=3)

my_str.set("output")

def add_data(): flag_validation = True #set the flag my_name=t1.get("1.0",END) #read name my_occupation=t2.get("1.0",END)#read occupation my_phonenumber=t3.get("1.0",END)#read phone number my_payment=radio_v.get()#read payment my_amount=t4.get("1.0",END)#read amount

#length of ny_name,my_class and my_gender more than 2
if len(my_name)<2 or len(my_occupation)<2 or len(my_payment)<2:
    flag_validation=False
try:
    val = int(my_payment)
    val=int(my_amount)          # checkinf phonenumber and amount as integers

except:
     flag_validation=False

 if flag_validation:
      try:
         #i=int(ws.cell(ws.max_row,1).value or 0)+1
         i=ws.max_row  #read the last row number
         j=ws.max_column
         if(i==1 and i!=j) or (i>1):
             i=i+1
             #print(str(ws.max_row) + ",i value :" +str(i))
          11=[i,my_name,my_occupation,my_phonenumber,my_payment,my_amount]   # row of data to add
          ws.append(11)

         wb.save(path)
         t1.delete("1.0",END) #reset the text entry box
         t2.delete("1.0",END) #reset text entry box
         t3.delete("1.0",END) #reset text entry box
         t4.delete("1.0",END) #reset text entry box
         l16.grid()
         l16.config(fg="green") # foreground color
         l16.config(bg="white") #background color
      #my_str.set("ID:"+str(id.lastrowid))   #this is for sql only
          my_str.set("sucess"+str(ws.max_row))
          l16.after(3000,lambda: l15.config(fg="yellow",bg="white",text="") )
       except Exception as e:
           print(e)

     else:
          l16.grid()
          l16.config(fg="red") #foreground color
          l16.config(bg="yellow") #background color
          my_str.set("check inputd.")
          l16.after(3000,lambda: l16.config(fg="white",bg="white",text=""))

my w.mainloop()
